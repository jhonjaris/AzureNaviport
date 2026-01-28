from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404, redirect
from solicitudes.models import Solicitud
from django import forms
from accounts.decorators import role_required
from django.contrib import messages
from django.db.models import Count, Q
from datetime import datetime, timedelta
from django.utils import timezone
from accounts.models import Empresa
from .models import ConfiguracionEvaluacion, Servicio, TipoLicencia, ConfiguracionEmail, DocumentoRequeridoServicio
from django.core.paginator import Paginator
from django.http import JsonResponse
from django.db.models import Q

# Create your views here.

@login_required
@role_required('evaluador')
def dashboard(request):
    # Obtener parámetros de filtro
    busqueda = request.GET.get('q', '')
    estado_filtro = request.GET.get('estado', '')
    prioridad_filtro = request.GET.get('prioridad', '')
    
    # Obtener todas las solicitudes base
    solicitudes_queryset = Solicitud.objects.all().select_related(
        'empresa', 'puerto_destino', 'motivo_acceso', 'solicitante'
    ).order_by('-creada_el')
    
    # Aplicar filtros de búsqueda
    if busqueda:
        solicitudes_queryset = solicitudes_queryset.filter(
            Q(empresa__nombre__icontains=busqueda) |
            Q(solicitante__first_name__icontains=busqueda) |
            Q(solicitante__last_name__icontains=busqueda) |
            Q(puerto_destino__nombre__icontains=busqueda) |
            Q(motivo_acceso__nombre__icontains=busqueda) |
            Q(id__icontains=busqueda)
        )
    
    # Aplicar filtro por estado
    if estado_filtro:
        if estado_filtro == 'pendiente':
            solicitudes_queryset = solicitudes_queryset.filter(estado__in=['pendiente', 'en_revision'])
        elif estado_filtro == 'sin_asignar':
            solicitudes_queryset = solicitudes_queryset.filter(estado='sin_asignar')
        elif estado_filtro == 'evaluadas':
            solicitudes_queryset = solicitudes_queryset.filter(estado__in=['aprobada', 'rechazada', 'documentos_requeridos'])
        elif estado_filtro == 'vencidas':
            solicitudes_queryset = solicitudes_queryset.filter(vence_el__lt=timezone.now(), estado__in=['pendiente', 'en_revision'])
        else:
            solicitudes_queryset = solicitudes_queryset.filter(estado=estado_filtro)
    
    # Aplicar filtro por prioridad
    if prioridad_filtro:
        solicitudes_queryset = solicitudes_queryset.filter(prioridad=prioridad_filtro)
    
    # Solicitudes pendientes (para estadísticas y tabla principal)
    solicitudes_pendientes = solicitudes_queryset.filter(
        estado__in=['pendiente', 'en_revision']
    )
    
    # Solicitudes asignadas a este evaluador
    mis_solicitudes = Solicitud.objects.filter(
        evaluador_asignado=request.user
    ).select_related('empresa', 'puerto_destino', 'motivo_acceso', 'solicitante').order_by('-creada_el')
    
    # Estadísticas del mes actual
    today = datetime.now().date()
    first_day_month = today.replace(day=1)
    
    # Estadísticas generales
    stats = {
        'pendientes_revision': solicitudes_pendientes.count(),
        'mis_asignadas': mis_solicitudes.filter(estado__in=['pendiente', 'en_revision']).count(),
        'vencidas_hoy': solicitudes_pendientes.filter(vence_el__date=today).count(),
        'criticas': solicitudes_pendientes.filter(prioridad__in=['critica', 'vip']).count(),
        
        # Estadísticas del mes
        'evaluadas_mes': Solicitud.objects.filter(
            evaluador_asignado=request.user,
            fecha_evaluacion__date__gte=first_day_month,
            estado__in=['aprobada', 'rechazada']
        ).count(),
        
        'aprobadas_mes': Solicitud.objects.filter(
            evaluador_asignado=request.user,
            fecha_evaluacion__date__gte=first_day_month,
            estado='aprobada'
        ).count(),
        
        'tiempo_promedio': 2.5,  # Simulado por ahora
    }
    
    # Calcular porcentaje de aprobación del mes
    if stats['evaluadas_mes'] > 0:
        stats['porcentaje_aprobacion'] = round((stats['aprobadas_mes'] / stats['evaluadas_mes']) * 100)
    else:
        stats['porcentaje_aprobacion'] = 0
    
    # Solicitudes por prioridad
    prioridades = solicitudes_pendientes.values('prioridad').annotate(
        count=Count('id')
    ).order_by('-count')
    
    # Paginación para las solicitudes filtradas
    per_page = request.GET.get('per_page', '25')
    try:
        per_page = int(per_page)
        if per_page not in [10, 25, 50, 100]:
            per_page = 25
    except (ValueError, TypeError):
        per_page = 25
    
    paginator = Paginator(solicitudes_queryset, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'user': request.user,
        'solicitudes': page_obj,  # Todas las solicitudes con filtros aplicados
        'solicitudes_pendientes': solicitudes_pendientes[:10] if not busqueda and not estado_filtro and not prioridad_filtro else [],  # Solo para el dashboard sin filtros
        'mis_solicitudes': mis_solicitudes[:5],  # Solo mostrar las primeras 5
        'stats': stats,
        'prioridades': prioridades,
        # Datos para filtros
        'busqueda': busqueda,
        'estado_filtro': estado_filtro,
        'prioridad_filtro': prioridad_filtro,
        'is_paginated': page_obj.has_other_pages(),
        'tiene_filtros': bool(busqueda or estado_filtro or prioridad_filtro)
    }
    
    return render(request, 'evaluacion/dashboard.html', context)

class EvaluacionForm(forms.Form):
    ACCIONES_CHOICES = [
        ('aprobar', '✅ Aprobar Solicitud'),
        ('rechazar', '❌ Rechazar Solicitud'),
        ('solicitar_documentos', '📄 Solicitar Documentos Adicionales'),
        ('escalar', '🚨 Escalar a Supervisor'),
    ]
    
    comentario = forms.CharField(
        widget=forms.Textarea(attrs={
            'rows': 4, 
            'class': 'form-control',
            'placeholder': 'Ingrese sus comentarios sobre la evaluación de esta solicitud...'
        }), 
        required=False, 
        label='Comentarios de Evaluación'
    )
    accion = forms.ChoiceField(
        choices=ACCIONES_CHOICES, 
        widget=forms.RadioSelect(attrs={'class': 'form-check-input'}), 
        label='Decisión'
    )

@login_required
@role_required('evaluador')
def evaluar_solicitud(request, solicitud_id):
    solicitud = get_object_or_404(Solicitud, id=solicitud_id)

    # Estados que permiten evaluación (incluye 'recibido' del wizard)
    ESTADOS_EVALUABLES = ['recibido', 'sin_asignar', 'pendiente', 'en_revision']

    # Verificar si la solicitud puede ser evaluada
    puede_evaluar = solicitud.estado in ESTADOS_EVALUABLES

    # Asignar el evaluador si no está asignado y puede ser evaluada
    if puede_evaluar and not solicitud.evaluador_asignado:
        solicitud.evaluador_asignado = request.user
        solicitud.estado = 'en_revision'
        solicitud.save()

    if request.method == 'POST' and puede_evaluar:
        form = EvaluacionForm(request.POST)
        if form.is_valid():
            accion = form.cleaned_data['accion']
            comentario = form.cleaned_data['comentario']
            
            # Actualizar la solicitud
            solicitud.fecha_evaluacion = datetime.now()
            solicitud.comentarios_evaluacion = comentario
            
            if accion == 'aprobar':
                solicitud.estado = 'aprobada'
                messages.success(request, "¡Solicitud aprobada exitosamente!")
                
                # Crear autorización automáticamente
                try:
                    from control_acceso.models import Autorizacion
                    autorizacion = Autorizacion.objects.create(
                        solicitud=solicitud,
                        empresa_nombre=solicitud.empresa.nombre,
                        empresa_rnc=solicitud.empresa.rnc,
                        representante_nombre=solicitud.solicitante.get_display_name(),
                        representante_cedula=solicitud.solicitante.cedula_rnc,
                        valida_desde=datetime.combine(solicitud.fecha_ingreso, solicitud.hora_ingreso),
                        valida_hasta=datetime.combine(solicitud.fecha_salida, solicitud.hora_salida),
                        puerto_nombre=solicitud.puerto_destino.nombre,
                        motivo_acceso=solicitud.motivo_acceso.nombre,
                        generada_por=request.user
                    )
                    messages.success(request, f"Autorización {autorizacion.codigo} generada automáticamente.")
                except Exception as e:
                    messages.warning(request, f"Solicitud aprobada pero error al generar autorización: {str(e)}")
                    
            elif accion == 'rechazar':
                solicitud.estado = 'rechazada'
                solicitud.motivo_rechazo = comentario
                messages.success(request, "¡Solicitud rechazada exitosamente!")
            elif accion == 'solicitar_documentos':
                solicitud.estado = 'documentos_faltantes'
                messages.success(request, "Se han solicitado documentos adicionales.")
            elif accion == 'escalar':
                solicitud.estado = 'escalada'
                # Crear escalamiento
                try:
                    from supervisor.models import Escalamiento
                    escalamiento = Escalamiento.objects.create(
                        solicitud=solicitud,
                        tipo_escalamiento='revision_manual',
                        prioridad='media',
                        motivo='Solicitud escalada por evaluador',
                        descripcion_detallada=comentario or 'Requiere revisión manual del supervisor',
                        escalado_por=request.user
                    )
                    messages.success(request, f"Solicitud escalada exitosamente. Código: {escalamiento.codigo}")
                except Exception as e:
                    messages.error(request, f"Error al crear escalamiento: {str(e)}")
            
            solicitud.save()
            return redirect('evaluacion:dashboard')
        else:
            messages.error(request, "Por favor corrige los errores en el formulario.")
    elif request.method == 'POST' and not puede_evaluar:
        messages.error(request, f"No se puede evaluar esta solicitud porque está en estado: {solicitud.get_estado_display()}")
    else:
        form = EvaluacionForm()

    # Obtener vehículos y documentos de la solicitud
    vehiculos = solicitud.vehiculos.all()
    documentos = solicitud.documentos.all()

    # Obtener personal asignado
    personal_asignado = solicitud.personal_asignado.select_related('personal').all()

    # Obtener servicios solicitados
    servicios = solicitud.servicios_solicitados.all()

    # Obtener documentos de servicios subidos
    from solicitudes.models import DocumentoServicioSolicitud, EventoSolicitud
    documentos_servicios = DocumentoServicioSolicitud.objects.filter(
        solicitud=solicitud
    ).select_related('documento_requerido', 'documento_requerido__servicio')

    # Obtener eventos del timeline
    eventos = solicitud.eventos.select_related('usuario').order_by('-creado_el')

    # Filtrar eventos según el rol del usuario
    if request.user.role in ['solicitante']:
        # Solicitantes solo ven eventos públicos
        eventos = eventos.filter(es_visible_solicitante=True)
    elif request.user.role in ['evaluador', 'supervisor', 'admin_tic', 'direccion']:
        # Staff ve todos los eventos
        pass

    # Serializar eventos para JavaScript
    import json
    eventos_json = json.dumps([{
        'id': evento.id,
        'tipo_evento': evento.tipo_evento,
        'titulo': evento.titulo,
        'descripcion': evento.descripcion,
        'usuario_nombre': evento.get_usuario_nombre(),
        'creado_el': evento.creado_el.isoformat(),
        'es_visible_solicitante': evento.es_visible_solicitante,
        'es_interno': evento.es_interno,
        'icono': evento.get_icono(),
        'color': evento.get_color(),
        'metadata': evento.metadata or {}
    } for evento in eventos])

    context = {
        'solicitud': solicitud,
        'form': form,
        'vehiculos': vehiculos,
        'documentos': documentos,
        'personal_asignado': personal_asignado,
        'servicios': servicios,
        'documentos_servicios': documentos_servicios,
        'puede_evaluar': puede_evaluar,
        'eventos': eventos,
        'eventos_json': eventos_json
    }
    return render(request, 'evaluacion/evaluar_solicitud.html', context)

# === GESTIÓN DE EMPRESAS ===

@login_required
@role_required('evaluador')
def gestionar_empresas(request):
    """Vista principal para gestionar empresas"""
    
    # Obtener parámetros de filtro y búsqueda
    busqueda = request.GET.get('q', '')
    estado_licencia = request.GET.get('estado', '')
    
    # Consulta base
    empresas = Empresa.objects.all().select_related('representante_legal').order_by('-created_at')
    
    # Aplicar filtros
    if busqueda:
        empresas = empresas.filter(
            Q(nombre__icontains=busqueda) |
            Q(rnc__icontains=busqueda) |
            Q(numero_licencia__icontains=busqueda) |
            Q(representante_legal__first_name__icontains=busqueda) |
            Q(representante_legal__last_name__icontains=busqueda)
        )
    
    # Filtro por estado de licencia
    if estado_licencia == 'vigente':
        empresas = empresas.filter(fecha_expiracion_licencia__gte=datetime.now().date())
    elif estado_licencia == 'proxima_vencer':
        proxima_fecha = datetime.now().date() + timedelta(days=30)
        empresas = empresas.filter(fecha_expiracion_licencia__lte=proxima_fecha, fecha_expiracion_licencia__gte=datetime.now().date())
    elif estado_licencia == 'vencida':
        empresas = empresas.filter(fecha_expiracion_licencia__lt=datetime.now().date())
    elif estado_licencia == 'sin_licencia':
        empresas = empresas.filter(Q(numero_licencia__isnull=True) | Q(numero_licencia=''))
    
    # Paginación
    paginator = Paginator(empresas, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Estadísticas para dashboard
    stats = {
        'total': Empresa.objects.count(),
        'vigentes': Empresa.objects.filter(fecha_expiracion_licencia__gte=datetime.now().date()).count(),
        'proximas_vencer': Empresa.objects.filter(
            fecha_expiracion_licencia__lte=datetime.now().date() + timedelta(days=30),
            fecha_expiracion_licencia__gte=datetime.now().date()
        ).count(),
        'vencidas': Empresa.objects.filter(fecha_expiracion_licencia__lt=datetime.now().date()).count(),
        'sin_licencia': Empresa.objects.filter(Q(numero_licencia__isnull=True) | Q(numero_licencia='')).count(),
    }
    
    context = {
        'empresas': page_obj,
        'busqueda': busqueda,
        'estado_licencia': estado_licencia,
        'stats': stats,
        'is_paginated': page_obj.has_other_pages(),
    }
    
    return render(request, 'evaluacion/gestionar_empresas.html', context)

class EmpresaForm(forms.ModelForm):
    """Formulario para crear/editar empresas"""
    
    servicios_autorizados = forms.ModelMultipleChoiceField(
        queryset=Servicio.objects.filter(activo=True).order_by('codigo'),
        widget=forms.CheckboxSelectMultiple(attrs={
            'class': 'servicios-checkboxes',
            'data-toggle': 'checkbox'
        }),
        required=False,
        help_text='Selecciona los servicios específicos que la empresa está autorizada a ofrecer'
    )
    
    class Meta:
        model = Empresa
        fields = ['rnc', 'nombre', 'telefono', 'email', 'numero_licencia', 
                 'fecha_expedicion_licencia', 'fecha_expiracion_licencia', 'fecha_expiracion_contrato', 
                 'tipo_licencia', 'servicios_autorizados', 'activa', 'verificada']
        widgets = {
            'rnc': forms.TextInput(attrs={
                'class': 'form-control', 
                'placeholder': 'XXX-XXXXX-X',
                'pattern': r'^\d{3}-\d{5}-\d{1}$'
            }),
            'nombre': forms.TextInput(attrs={'class': 'form-control'}),
            'telefono': forms.TextInput(attrs={'class': 'form-control', 'placeholder': '+1-XXX-XXX-XXXX'}),
            'email': forms.EmailInput(attrs={'class': 'form-control'}),
            'numero_licencia': forms.TextInput(attrs={'class': 'form-control'}),
            'fecha_expedicion_licencia': forms.DateInput(attrs={'class': 'form-control', 'type': 'date'}),
            'fecha_expiracion_licencia': forms.DateInput(attrs={'class': 'form-control', 'type': 'date'}),
            'fecha_expiracion_contrato': forms.DateInput(attrs={'class': 'form-control', 'type': 'date'}),
            'tipo_licencia': forms.Select(attrs={
                'class': 'form-control',
                'id': 'id_tipo_licencia'
            }),
            'activa': forms.CheckboxInput(attrs={'class': 'form-check-input'}),
            'verificada': forms.CheckboxInput(attrs={'class': 'form-check-input'}),
        }
        
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        
        # Configurar queryset para tipo_licencia
        self.fields['tipo_licencia'].queryset = TipoLicencia.objects.filter(activo=True).order_by('nombre')
        self.fields['tipo_licencia'].empty_label = "--- Seleccionar Tipo de Licencia ---"
        
        # Si estamos editando una empresa existente, marcar los servicios actuales
        if self.instance and self.instance.pk:
            self.initial['servicios_autorizados'] = self.instance.servicios_autorizados.all()
    
    def save(self, commit=True):
        empresa = super().save(commit=False)
        if commit:
            empresa.save()
            # Guardar servicios autorizados (ManyToMany)
            self.save_m2m()
        return empresa

@login_required
@role_required('evaluador')
def crear_empresa(request):
    """Vista para crear nueva empresa"""
    if request.method == 'POST':
        form = EmpresaForm(request.POST)
        if form.is_valid():
            empresa = form.save(commit=False)
            # Asignar representante legal temporal (esto debería mejorarse)
            empresa.representante_legal = request.user
            empresa.save()
            # Guardar los servicios autorizados (ManyToMany)
            form.save_m2m()
            messages.success(request, f'Empresa "{empresa.nombre}" creada exitosamente.')
            return redirect('evaluacion:gestionar_empresas')
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = EmpresaForm()
    
    context = {
        'form': form,
        'tipos_licencia': TipoLicencia.objects.filter(activo=True).order_by('nombre'),
        'servicios': Servicio.objects.filter(activo=True).order_by('codigo')
    }
    return render(request, 'evaluacion/crear_empresa.html', context)

@login_required
@role_required('evaluador')
def editar_empresa(request, empresa_id):
    """Vista para editar empresa existente"""
    empresa = get_object_or_404(Empresa, id=empresa_id)
    
    if request.method == 'POST':
        form = EmpresaForm(request.POST, instance=empresa)
        if form.is_valid():
            form.save()
            messages.success(request, f'Empresa "{empresa.nombre}" actualizada exitosamente.')
            return redirect('evaluacion:gestionar_empresas')
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = EmpresaForm(instance=empresa)
    
    context = {
        'form': form,
        'empresa': empresa,
        'editando': True,
        'tipos_licencia': TipoLicencia.objects.filter(activo=True).order_by('nombre'),
        'servicios': Servicio.objects.filter(activo=True).order_by('codigo')
    }
    return render(request, 'evaluacion/crear_empresa.html', context)

@login_required
@role_required('evaluador')
def renovar_licencia(request, empresa_id):
    """Vista para renovar licencia de empresa"""
    empresa = get_object_or_404(Empresa, id=empresa_id)
    
    if request.method == 'POST':
        nueva_fecha = request.POST.get('fecha_expiracion')
        if nueva_fecha:
            try:
                empresa.fecha_expiracion_licencia = datetime.strptime(nueva_fecha, '%Y-%m-%d').date()
                empresa.save()
                messages.success(request, f'Licencia de "{empresa.nombre}" renovada hasta {nueva_fecha}.')
            except ValueError:
                messages.error(request, 'Fecha inválida.')
        else:
            messages.error(request, 'Debe especificar una fecha de expiración.')
        
        return redirect('evaluacion:gestionar_empresas')
    
    # Sugerir fecha de renovación (1 año desde hoy)
    fecha_sugerida = (datetime.now().date() + timedelta(days=365)).strftime('%Y-%m-%d')
    
    context = {
        'empresa': empresa,
        'fecha_sugerida': fecha_sugerida
    }
    return render(request, 'evaluacion/renovar_licencia.html', context)

@login_required
@role_required('evaluador')
def eliminar_empresa(request, empresa_id):
    """Vista para eliminar empresa"""
    empresa = get_object_or_404(Empresa, id=empresa_id)
    
    if request.method == 'POST':
        nombre_empresa = empresa.nombre
        empresa.delete()
        messages.success(request, f'Empresa "{nombre_empresa}" eliminada exitosamente.')
        return redirect('evaluacion:gestionar_empresas')
    
    # Verificar si la empresa tiene solicitudes asociadas
    solicitudes_count = empresa.solicitudes.count()
    
    context = {
        'empresa': empresa,
        'solicitudes_count': solicitudes_count
    }
    return render(request, 'evaluacion/eliminar_empresa.html', context)

@login_required
@role_required('evaluador')  
def buscar_empresas_ajax(request):
    """Vista AJAX para búsqueda en tiempo real"""
    query = request.GET.get('q', '')
    
    if len(query) < 2:
        return JsonResponse({'results': []})
    
    empresas = Empresa.objects.filter(
        Q(nombre__icontains=query) |
        Q(rnc__icontains=query) |
        Q(numero_licencia__icontains=query)
    )[:10]
    
    results = []
    for empresa in empresas:
        estado = empresa.estado_licencia
        results.append({
            'id': empresa.id,
            'nombre': empresa.nombre,
            'rnc': empresa.rnc,
            'numero_licencia': empresa.numero_licencia or 'Sin licencia',
            'fecha_expiracion': empresa.fecha_expiracion_licencia.strftime('%d/%m/%Y') if empresa.fecha_expiracion_licencia else 'N/A',
            'estado_licencia': estado['texto'],
            'color_estado': estado['color']
        })
    
    return JsonResponse({'results': results})

@login_required
@role_required('evaluador')
def obtener_servicios_tipo_licencia(request, tipo_licencia_id):
    """Vista AJAX para obtener servicios incluidos en un tipo de licencia"""
    try:
        tipo_licencia = get_object_or_404(TipoLicencia, id=tipo_licencia_id, activo=True)
        servicios = tipo_licencia.servicios_incluidos.filter(activo=True).values('id', 'codigo', 'nombre')
        
        return JsonResponse({
            'servicios': list(servicios),
            'tipo_licencia': tipo_licencia.nombre
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

class ConfiguracionEvaluacionForm(forms.ModelForm):
    """Formulario para configurar las reglas de evaluación"""
    
    class Meta:
        model = ConfiguracionEvaluacion
        fields = [
            'dias_preaviso_critico', 'dias_preaviso_advertencia', 'dias_preaviso_informativo',
            'tipo_expiracion_principal', 'enlace_instrucciones', 'tiempo_respuesta_horas', 'notificar_empresas_vencidas'
        ]
        widgets = {
            'dias_preaviso_critico': forms.NumberInput(attrs={
                'class': 'form-control',
                'min': '1',
                'max': '365',
                'placeholder': '30'
            }),
            'dias_preaviso_advertencia': forms.NumberInput(attrs={
                'class': 'form-control',
                'min': '1',
                'max': '365', 
                'placeholder': '60'
            }),
            'dias_preaviso_informativo': forms.NumberInput(attrs={
                'class': 'form-control',
                'min': '1',
                'max': '365',
                'placeholder': '90'
            }),
            'tipo_expiracion_principal': forms.Select(attrs={'class': 'form-control'}),
            'enlace_instrucciones': forms.URLInput(attrs={
                'class': 'form-control',
                'placeholder': 'https://vuce.gob.do/instrucciones-renovacion'
            }),
            'tiempo_respuesta_horas': forms.NumberInput(attrs={
                'class': 'form-control',
                'min': '1',
                'max': '168',
                'placeholder': '24'
            }),
            'notificar_empresas_vencidas': forms.CheckboxInput(attrs={'class': 'form-check-input'}),
        }

@login_required
@role_required('evaluador')
def configuracion(request):
    """Vista para configurar las reglas de evaluación"""
    config = ConfiguracionEvaluacion.get_configuracion()
    
    if request.method == 'POST':
        form = ConfiguracionEvaluacionForm(request.POST, instance=config)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, 'Configuración actualizada correctamente.')
                return redirect('evaluacion:configuracion')
            except Exception as e:
                messages.error(request, f'Error al guardar la configuración: {str(e)}')
        else:
            messages.error(request, 'Por favor corrija los errores en el formulario.')
    else:
        form = ConfiguracionEvaluacionForm(instance=config)
    
    context = {
        'form': form,
        'config': config,
    }
    
    return render(request, 'evaluacion/configuracion.html', context)

# === GESTIÓN DE SERVICIOS ===

@login_required
@role_required('evaluador')
def gestionar_servicios(request):
    """Vista principal para gestionar servicios"""
    
    # Obtener parámetros de filtro y búsqueda
    busqueda = request.GET.get('q', '')
    estado_filtro = request.GET.get('estado', '')
    
    # Consulta base
    servicios = Servicio.objects.all().order_by('codigo', 'nombre')
    
    # Aplicar filtros
    if busqueda:
        servicios = servicios.filter(
            Q(nombre__icontains=busqueda) |
            Q(codigo__icontains=busqueda) |
            Q(descripcion__icontains=busqueda)
        )
    
    if estado_filtro == 'activo':
        servicios = servicios.filter(activo=True)
    elif estado_filtro == 'inactivo':
        servicios = servicios.filter(activo=False)
    
    # Paginación
    paginator = Paginator(servicios, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Estadísticas
    stats = {
        'total': Servicio.objects.count(),
        'activos': Servicio.objects.filter(activo=True).count(),
        'inactivos': Servicio.objects.filter(activo=False).count(),
    }
    
    context = {
        'servicios': page_obj,
        'busqueda': busqueda,
        'estado_filtro': estado_filtro,
        'stats': stats,
        'is_paginated': page_obj.has_other_pages(),
    }
    
    return render(request, 'evaluacion/gestionar_servicios.html', context)


class ServicioForm(forms.ModelForm):
    """Formulario para crear/editar servicios"""

    class Meta:
        model = Servicio
        fields = ['codigo', 'nombre', 'descripcion', 'activo']
        widgets = {
            'codigo': forms.TextInput(attrs={
                'class': 'form-control',
                'placeholder': 'CG-001, CNT-002, etc.'
            }),
            'nombre': forms.TextInput(attrs={'class': 'form-control'}),
            'descripcion': forms.Textarea(attrs={
                'class': 'form-control',
                'rows': 3,
                'placeholder': 'Descripción detallada del servicio...'
            }),
            'activo': forms.CheckboxInput(attrs={'class': 'form-check-input'}),
        }


class DocumentoRequeridoForm(forms.ModelForm):
    """Formulario para documentos requeridos de un servicio"""

    class Meta:
        model = DocumentoRequeridoServicio
        fields = ['nombre', 'descripcion', 'obligatorio', 'orden', 'activo']
        widgets = {
            'nombre': forms.TextInput(attrs={
                'class': 'form-control',
                'placeholder': 'Nombre del documento requerido'
            }),
            'descripcion': forms.Textarea(attrs={
                'class': 'form-control',
                'rows': 2,
                'placeholder': 'Instrucciones para el solicitante...'
            }),
            'obligatorio': forms.CheckboxInput(attrs={'class': 'form-check-input'}),
            'orden': forms.NumberInput(attrs={'class': 'form-control', 'min': 0, 'style': 'width: 80px;'}),
            'activo': forms.CheckboxInput(attrs={'class': 'form-check-input'}),
        }


# Formset para documentos requeridos (máximo 5)
DocumentoRequeridoFormSet = forms.inlineformset_factory(
    Servicio,
    DocumentoRequeridoServicio,
    form=DocumentoRequeridoForm,
    extra=1,
    max_num=5,
    validate_max=True,
    can_delete=True
)


@login_required
@role_required('evaluador')
def crear_servicio(request):
    """Vista para crear nuevo servicio"""
    if request.method == 'POST':
        form = ServicioForm(request.POST)
        formset = DocumentoRequeridoFormSet(request.POST, prefix='documentos')

        if form.is_valid() and formset.is_valid():
            servicio = form.save()
            # Guardar documentos requeridos
            documentos = formset.save(commit=False)
            for doc in documentos:
                doc.servicio = servicio
                doc.save()
            # Eliminar documentos marcados para borrar
            for obj in formset.deleted_objects:
                obj.delete()
            messages.success(request, f'Servicio "{servicio.nombre}" creado exitosamente.')
            return redirect('evaluacion:gestionar_servicios')
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = ServicioForm()
        formset = DocumentoRequeridoFormSet(prefix='documentos')

    return render(request, 'evaluacion/crear_servicio.html', {
        'form': form,
        'formset': formset,
    })


@login_required
@role_required('evaluador')
def editar_servicio(request, servicio_id):
    """Vista para editar servicio existente"""
    servicio = get_object_or_404(Servicio, id=servicio_id)

    if request.method == 'POST':
        form = ServicioForm(request.POST, instance=servicio)
        formset = DocumentoRequeridoFormSet(request.POST, instance=servicio, prefix='documentos')

        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()
            messages.success(request, f'Servicio "{servicio.nombre}" actualizado exitosamente.')
            return redirect('evaluacion:gestionar_servicios')
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = ServicioForm(instance=servicio)
        formset = DocumentoRequeridoFormSet(instance=servicio, prefix='documentos')

    context = {
        'form': form,
        'formset': formset,
        'servicio': servicio,
        'editando': True
    }
    return render(request, 'evaluacion/crear_servicio.html', context)


@login_required
@role_required('evaluador')
def eliminar_servicio(request, servicio_id):
    """Vista para eliminar servicio"""
    servicio = get_object_or_404(Servicio, id=servicio_id)
    
    if request.method == 'POST':
        if servicio.puede_eliminar():
            nombre_servicio = servicio.nombre
            servicio.delete()
            messages.success(request, f'Servicio "{nombre_servicio}" eliminado exitosamente.')
        else:
            # Mostrar empresas y tipos de licencia que lo usan
            empresas = list(servicio.empresas_que_lo_usan())
            tipos_licencia = list(servicio.tipos_licencia_que_lo_incluyen())
            
            mensaje_error = f'No se puede eliminar el servicio "{servicio.nombre}" porque está siendo utilizado por:'
            if empresas:
                mensaje_error += f' {len(empresas)} empresa(s)'
            if tipos_licencia:
                mensaje_error += f' {len(tipos_licencia)} tipo(s) de licencia'
                
            messages.error(request, mensaje_error)
        
        return redirect('evaluacion:gestionar_servicios')
    
    # Obtener información de uso
    empresas_usando = servicio.empresas_que_lo_usan()
    tipos_licencia_usando = servicio.tipos_licencia_que_lo_incluyen()
    puede_eliminar = servicio.puede_eliminar()
    
    context = {
        'servicio': servicio,
        'empresas_usando': empresas_usando,
        'tipos_licencia_usando': tipos_licencia_usando,
        'puede_eliminar': puede_eliminar
    }
    return render(request, 'evaluacion/eliminar_servicio.html', context)


# === GESTIÓN DE TIPOS DE LICENCIA ===

@login_required
@role_required('evaluador')
def gestionar_tipos_licencia(request):
    """Vista principal para gestionar tipos de licencia"""

    # Obtener parámetros de filtro y búsqueda
    busqueda = request.GET.get('q', '')
    estado_filtro = request.GET.get('estado', '')
    servicio_filtro = request.GET.get('servicio', '')

    # Consulta base
    tipos_licencia = TipoLicencia.objects.all().order_by('nombre')

    # Aplicar filtros
    if busqueda:
        tipos_licencia = tipos_licencia.filter(
            Q(nombre__icontains=busqueda) |
            Q(descripcion__icontains=busqueda)
        )

    if estado_filtro == 'activo':
        tipos_licencia = tipos_licencia.filter(activo=True)
    elif estado_filtro == 'inactivo':
        tipos_licencia = tipos_licencia.filter(activo=False)

    # Filtro por servicio incluido
    if servicio_filtro:
        try:
            servicio_id = int(servicio_filtro)
            tipos_licencia = tipos_licencia.filter(servicios_incluidos__id=servicio_id)
        except (ValueError, TypeError):
            pass

    # Paginación
    paginator = Paginator(tipos_licencia, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Estadísticas
    stats = {
        'total': TipoLicencia.objects.count(),
        'activos': TipoLicencia.objects.filter(activo=True).count(),
        'inactivos': TipoLicencia.objects.filter(activo=False).count(),
    }

    # Lista de servicios para el filtro
    servicios_disponibles = Servicio.objects.filter(activo=True).order_by('nombre')

    context = {
        'tipos_licencia': page_obj,
        'busqueda': busqueda,
        'estado_filtro': estado_filtro,
        'servicio_filtro': servicio_filtro,
        'servicios_disponibles': servicios_disponibles,
        'stats': stats,
        'is_paginated': page_obj.has_other_pages(),
    }

    return render(request, 'evaluacion/gestionar_tipos_licencia.html', context)


class TipoLicenciaForm(forms.ModelForm):
    """Formulario para crear/editar tipos de licencia"""
    
    servicios_incluidos = forms.ModelMultipleChoiceField(
        queryset=Servicio.objects.filter(activo=True),
        widget=forms.CheckboxSelectMultiple,
        required=False,
        help_text='Selecciona los servicios que se incluyen automáticamente con este tipo de licencia'
    )
    
    class Meta:
        model = TipoLicencia
        fields = ['nombre', 'descripcion', 'servicios_incluidos', 'activo']
        widgets = {
            'nombre': forms.TextInput(attrs={
                'class': 'form-control',
                'placeholder': 'Ej: Licencia Portuaria General'
            }),
            'descripcion': forms.Textarea(attrs={
                'class': 'form-control',
                'rows': 3,
                'placeholder': 'Descripción del tipo de licencia y sus alcances...'
            }),
            'activo': forms.CheckboxInput(attrs={'class': 'form-check-input'}),
        }


@login_required
@role_required('evaluador')
def crear_tipo_licencia(request):
    """Vista para crear nuevo tipo de licencia"""
    if request.method == 'POST':
        form = TipoLicenciaForm(request.POST)
        if form.is_valid():
            tipo_licencia = form.save()
            messages.success(request, f'Tipo de licencia "{tipo_licencia.nombre}" creado exitosamente.')
            return redirect('evaluacion:gestionar_tipos_licencia')
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = TipoLicenciaForm()
    
    return render(request, 'evaluacion/crear_tipo_licencia.html', {'form': form})


@login_required
@role_required('evaluador')
def editar_tipo_licencia(request, tipo_licencia_id):
    """Vista para editar tipo de licencia existente"""
    tipo_licencia = get_object_or_404(TipoLicencia, id=tipo_licencia_id)
    
    if request.method == 'POST':
        form = TipoLicenciaForm(request.POST, instance=tipo_licencia)
        if form.is_valid():
            form.save()
            messages.success(request, f'Tipo de licencia "{tipo_licencia.nombre}" actualizado exitosamente.')
            return redirect('evaluacion:gestionar_tipos_licencia')
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = TipoLicenciaForm(instance=tipo_licencia)
    
    context = {
        'form': form,
        'tipo_licencia': tipo_licencia,
        'editando': True
    }
    return render(request, 'evaluacion/crear_tipo_licencia.html', context)


@login_required
@role_required('evaluador')
def eliminar_tipo_licencia(request, tipo_licencia_id):
    """Vista para eliminar tipo de licencia"""
    tipo_licencia = get_object_or_404(TipoLicencia, id=tipo_licencia_id)
    
    if request.method == 'POST':
        if tipo_licencia.puede_eliminar():
            nombre_tipo = tipo_licencia.nombre
            tipo_licencia.delete()
            messages.success(request, f'Tipo de licencia "{nombre_tipo}" eliminado exitosamente.')
        else:
            # Mostrar empresas que lo usan
            empresas = list(tipo_licencia.empresas_que_lo_usan())
            mensaje_error = f'No se puede eliminar el tipo de licencia "{tipo_licencia.nombre}" porque está siendo utilizado por {len(empresas)} empresa(s).'
            messages.error(request, mensaje_error)
        
        return redirect('evaluacion:gestionar_tipos_licencia')
    
    # Obtener información de uso
    empresas_usando = tipo_licencia.empresas_que_lo_usan()
    puede_eliminar = tipo_licencia.puede_eliminar()
    
    context = {
        'tipo_licencia': tipo_licencia,
        'empresas_usando': empresas_usando,
        'puede_eliminar': puede_eliminar
    }
    return render(request, 'evaluacion/eliminar_tipo_licencia.html', context)


@login_required
@role_required('evaluador')
def ver_tipo_licencia(request, tipo_licencia_id):
    """Vista para ver detalles del tipo de licencia"""
    tipo_licencia = get_object_or_404(TipoLicencia, id=tipo_licencia_id)
    servicios_incluidos = tipo_licencia.servicios_incluidos.filter(activo=True)
    empresas_asignadas = tipo_licencia.empresas_que_lo_usan()
    
    context = {
        'tipo_licencia': tipo_licencia,
        'servicios_incluidos': servicios_incluidos,
        'empresas_asignadas': empresas_asignadas,
        'cantidad_servicios': servicios_incluidos.count(),
        'cantidad_empresas': empresas_asignadas.count()
    }
    return render(request, 'evaluacion/ver_tipo_licencia.html', context)


@login_required
@role_required('evaluador')
def gestion_licencias_servicios(request):
    """Vista principal para gestión integrada de licencias y servicios"""
    
    # Estadísticas de servicios
    stats_servicios = {
        'total': Servicio.objects.count(),
        'activos': Servicio.objects.filter(activo=True).count(),
        'inactivos': Servicio.objects.filter(activo=False).count(),
    }
    
    # Estadísticas de tipos de licencia
    stats_tipos_licencia = {
        'total': TipoLicencia.objects.count(),
        'activos': TipoLicencia.objects.filter(activo=True).count(),
        'inactivos': TipoLicencia.objects.filter(activo=False).count(),
    }
    
    # Servicios más utilizados
    servicios_populares = Servicio.objects.filter(activo=True).annotate(
        empresas_count=Count('empresas_autorizadas')
    ).order_by('-empresas_count')[:5]
    
    # Tipos de licencia más utilizados
    tipos_populares = TipoLicencia.objects.filter(activo=True).annotate(
        empresas_count=Count('empresas_asignadas')
    ).order_by('-empresas_count')[:5]
    
    # Servicios recientes
    servicios_recientes = Servicio.objects.filter(activo=True).order_by('-created_at')[:5]
    
    # Tipos de licencia recientes
    tipos_recientes = TipoLicencia.objects.filter(activo=True).order_by('-created_at')[:5]
    
    context = {
        'user': request.user,
        'stats_servicios': stats_servicios,
        'stats_tipos_licencia': stats_tipos_licencia,
        'servicios_populares': servicios_populares,
        'tipos_populares': tipos_populares,
        'servicios_recientes': servicios_recientes,
        'tipos_recientes': tipos_recientes,
    }
    
    return render(request, 'evaluacion/gestion_licencias_servicios.html', context)

@login_required
@role_required('evaluador')
def mis_solicitudes(request):
    """Vista para mostrar las solicitudes asignadas al evaluador actual"""
    from accounts.models import User
    
    # Obtener parámetros de filtro
    busqueda = request.GET.get('q', '')
    estado_filtro = request.GET.get('estado', '')
    prioridad_filtro = request.GET.get('prioridad', '')
    
    # Obtener solicitudes asignadas al usuario actual
    solicitudes_queryset = Solicitud.objects.filter(
        evaluador_asignado=request.user
    ).select_related(
        'empresa', 'puerto_destino', 'motivo_acceso', 'solicitante'
    ).order_by('-creada_el')
    
    # Aplicar filtros de búsqueda
    if busqueda:
        solicitudes_queryset = solicitudes_queryset.filter(
            Q(empresa__nombre__icontains=busqueda) |
            Q(solicitante__first_name__icontains=busqueda) |
            Q(solicitante__last_name__icontains=busqueda) |
            Q(puerto_destino__nombre__icontains=busqueda) |
            Q(motivo_acceso__nombre__icontains=busqueda) |
            Q(id__icontains=busqueda)
        )
    
    # Aplicar filtro por estado
    if estado_filtro:
        solicitudes_queryset = solicitudes_queryset.filter(estado=estado_filtro)
    
    # Aplicar filtro por prioridad
    if prioridad_filtro:
        solicitudes_queryset = solicitudes_queryset.filter(prioridad=prioridad_filtro)
    
    # Estadísticas específicas del evaluador
    stats = {
        'total_asignadas': Solicitud.objects.filter(evaluador_asignado=request.user).count(),
        'en_proceso': solicitudes_queryset.filter(estado__in=['pendiente', 'en_revision']).count(),
        'completadas': solicitudes_queryset.filter(estado__in=['aprobada', 'rechazada']).count(),
        'documentos_pendientes': solicitudes_queryset.filter(estado='documentos_faltantes').count(),
    }
    
    # Paginación
    per_page = request.GET.get('per_page', '25')
    try:
        per_page = int(per_page)
        if per_page not in [10, 25, 50, 100]:
            per_page = 25
    except (ValueError, TypeError):
        per_page = 25
    
    paginator = Paginator(solicitudes_queryset, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Verificar si hay filtros aplicados
    tiene_filtros = bool(busqueda or estado_filtro or prioridad_filtro)
    
    context = {
        'user': request.user,
        'solicitudes': page_obj,
        'stats': stats,
        'busqueda': busqueda,
        'estado_filtro': estado_filtro,
        'prioridad_filtro': prioridad_filtro,
        'tiene_filtros': tiene_filtros,
        'is_paginated': page_obj.has_other_pages(),
    }
    
    return render(request, 'evaluacion/mis_solicitudes.html', context)

@login_required
@role_required('evaluador')
def nuevas_solicitudes(request):
    """Vista para mostrar las solicitudes nuevas (sin asignar)"""
    from accounts.models import User
    
    # Obtener parámetros de filtro
    busqueda = request.GET.get('q', '')
    prioridad_filtro = request.GET.get('prioridad', '')
    
    # Obtener solicitudes sin asignar
    solicitudes_queryset = Solicitud.objects.filter(
        estado='sin_asignar'
    ).select_related(
        'empresa', 'puerto_destino', 'motivo_acceso', 'solicitante'
    ).order_by('-creada_el')
    
    # Aplicar filtros de búsqueda
    if busqueda:
        solicitudes_queryset = solicitudes_queryset.filter(
            Q(empresa__nombre__icontains=busqueda) |
            Q(solicitante__first_name__icontains=busqueda) |
            Q(solicitante__last_name__icontains=busqueda) |
            Q(puerto_destino__nombre__icontains=busqueda) |
            Q(motivo_acceso__nombre__icontains=busqueda) |
            Q(id__icontains=busqueda)
        )
    
    # Aplicar filtro por prioridad
    if prioridad_filtro:
        solicitudes_queryset = solicitudes_queryset.filter(prioridad=prioridad_filtro)
    
    # Estadísticas de solicitudes nuevas
    stats = {
        'total_nuevas': Solicitud.objects.filter(estado='sin_asignar').count(),
        'alta_prioridad': solicitudes_queryset.filter(prioridad__in=['alta', 'critica', 'vip']).count(),
        'vencen_pronto': solicitudes_queryset.filter(
            vence_el__lte=timezone.now() + timedelta(hours=24)
        ).count() if solicitudes_queryset.exists() else 0,
        'hoy': solicitudes_queryset.filter(creada_el__date=timezone.now().date()).count(),
    }
    
    # Paginación
    per_page = request.GET.get('per_page', '25')
    try:
        per_page = int(per_page)
        if per_page not in [10, 25, 50, 100]:
            per_page = 25
    except (ValueError, TypeError):
        per_page = 25
    
    paginator = Paginator(solicitudes_queryset, per_page)
    page_number = request.GET.get('page')
    solicitudes = paginator.get_page(page_number)
    
    # Detectar si hay filtros aplicados
    tiene_filtros = bool(busqueda or prioridad_filtro)
    
    context = {
        'solicitudes': solicitudes,
        'stats': stats,
        'busqueda': busqueda,
        'prioridad_filtro': prioridad_filtro,
        'tiene_filtros': tiene_filtros,
        'is_paginated': paginator.num_pages > 1,
    }
    
    return render(request, 'evaluacion/nuevas_solicitudes.html', context)

@login_required
@role_required('evaluador')
def asignar_evaluador(request, solicitud_id):
    """Vista para asignar un evaluador a una solicitud"""
    from accounts.models import User

    solicitud = get_object_or_404(Solicitud, id=solicitud_id)

    if request.method == 'POST':
        evaluador_id = request.POST.get('evaluador_id')
        if evaluador_id:
            try:
                evaluador = User.objects.get(id=evaluador_id, role='evaluador')
                solicitud.evaluador_asignado = evaluador
                # Cambiar de 'sin_asignar' a 'pendiente' cuando se asigna evaluador
                solicitud.estado = 'pendiente'
                solicitud.save()

                messages.success(request, f'Solicitud #{solicitud.id} asignada exitosamente a {evaluador.get_display_name()}')
                return JsonResponse({'success': True, 'message': 'Asignación exitosa'})
            except User.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Evaluador no encontrado'})
        else:
            return JsonResponse({'success': False, 'error': 'Debe seleccionar un evaluador'})

    # GET request - obtener evaluadores disponibles
    evaluadores = User.objects.filter(role='evaluador', is_active=True).exclude(id=request.user.id)

    return JsonResponse({
        'usuario_actual': {
            'id': request.user.id,
            'nombre': request.user.get_display_name(),
            'solicitudes_asignadas': request.user.solicitudes_asignadas.filter(estado='en_revision').count()
        },
        'evaluadores': [
            {
                'id': eval.id,
                'nombre': eval.get_display_name(),
                'solicitudes_asignadas': eval.solicitudes_asignadas.filter(estado='en_revision').count()
            }
            for eval in evaluadores
        ]
    })


# === CONFIGURACIÓN DE EMAIL ===

class ConfiguracionEmailForm(forms.ModelForm):
    """Formulario para configurar el sistema de email"""
    
    tipo_proveedor = forms.ChoiceField(
        choices=ConfiguracionEmail.TIPO_PROVEEDOR_CHOICES,
        widget=forms.Select(attrs={'class': 'form-control'})
    )
    
    email_host_password = forms.CharField(
        widget=forms.PasswordInput(attrs={
            'class': 'form-control',
            'placeholder': '•••••••••••••'
        }),
        label='Contraseña de Email',
        help_text='Para Gmail usa App Password en lugar de tu contraseña normal',
        required=False
    )
    
    class Meta:
        model = ConfiguracionEmail
        fields = [
            'nombre', 'tipo_proveedor', 'email_host', 'email_port', 'email_use_tls', 'email_use_ssl',
            'email_host_user', 'email_host_password', 'default_from_email',
            'password_reset_timeout', 'max_reset_attempts', 'reset_email_subject',
            'correos_notificaciones', 'email_enabled'
        ]
        widgets = {
            'nombre': forms.TextInput(attrs={
                'class': 'form-control',
                'placeholder': 'Ej: Gmail Principal'
            }),
            'email_host': forms.TextInput(attrs={
                'class': 'form-control',
                'placeholder': 'smtp.gmail.com'
            }),
            'email_port': forms.NumberInput(attrs={
                'class': 'form-control',
                'placeholder': '587'
            }),
            'email_use_tls': forms.CheckboxInput(attrs={'class': 'form-check-input'}),
            'email_use_ssl': forms.CheckboxInput(attrs={'class': 'form-check-input'}),
            'email_host_user': forms.EmailInput(attrs={
                'class': 'form-control',
                'placeholder': 'tu-email@gmail.com'
            }),
            'default_from_email': forms.EmailInput(attrs={
                'class': 'form-control',
                'placeholder': 'NaviPort RD <noreply@naviport.com>'
            }),
            'password_reset_timeout': forms.NumberInput(attrs={
                'class': 'form-control',
                'min': '1',
                'max': '168'
            }),
            'max_reset_attempts': forms.NumberInput(attrs={
                'class': 'form-control',
                'min': '1',
                'max': '10'
            }),
            'reset_email_subject': forms.TextInput(attrs={
                'class': 'form-control',
                'placeholder': 'NaviPort RD - Recuperación de Contraseña'
            }),
            'correos_notificaciones': forms.Textarea(attrs={
                'class': 'form-control',
                'rows': '3',
                'placeholder': 'evaluador1@ejemplo.com, evaluador2@ejemplo.com, supervisor@ejemplo.com'
            }),
            'email_enabled': forms.CheckboxInput(attrs={'class': 'form-check-input'}),
        }
    
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        # Guardar el valor original de la contraseña
        if self.instance and self.instance.pk:
            self.existing_password = self.instance.email_host_password
        else:
            self.existing_password = None
    
    def clean_email_host_password(self):
        password = self.cleaned_data.get('email_host_password')
        if self.instance and self.instance.pk and not password:
            return self.existing_password
        return password
    
    def clean(self):
        cleaned_data = super().clean()
        usar_tls = cleaned_data.get('email_use_tls')
        usar_ssl = cleaned_data.get('email_use_ssl')
        tipo_proveedor = cleaned_data.get('tipo_proveedor')

        if usar_tls and usar_ssl:
            raise forms.ValidationError('No puede usar TLS y SSL simultáneamente')

        # Configuración automática según el proveedor
        if tipo_proveedor == 'gmail':
            cleaned_data['email_host'] = 'smtp.gmail.com'
            cleaned_data['email_port'] = 587
            cleaned_data['email_use_tls'] = True
            cleaned_data['email_use_ssl'] = False
        elif tipo_proveedor == 'outlook':
            cleaned_data['email_host'] = 'smtp.office365.com'
            cleaned_data['email_port'] = 587
            cleaned_data['email_use_tls'] = True
            cleaned_data['email_use_ssl'] = False

        return cleaned_data
    
    def save(self, commit=True):
        instance = super().save(commit=False)
        
        # Configurar valores según el proveedor
        if instance.tipo_proveedor == 'gmail':
            instance.email_host = 'smtp.gmail.com'
            instance.email_port = 587
            instance.email_use_tls = True
            instance.email_use_ssl = False
        elif instance.tipo_proveedor == 'outlook':
            instance.email_host = 'smtp.office365.com'
            instance.email_port = 587
            instance.email_use_tls = True
            instance.email_use_ssl = False
        
        # Si esta configuración está activa, desactivar las demás
        if instance.email_enabled:
            ConfiguracionEmail.objects.exclude(pk=instance.pk).update(email_enabled=False)
        
        if commit:
            instance.save()
        return instance

@login_required
@role_required('evaluador')
def configuracion_email(request):
    """Vista principal para gestionar configuraciones de correo"""
    configuraciones = ConfiguracionEmail.objects.all().order_by('-email_enabled', '-updated_at')
    configuracion_activa = ConfiguracionEmail.objects.filter(email_enabled=True).first()
    
    return render(request, 'evaluacion/configuracion_email_lista.html', {
        'configuraciones': configuraciones,
        'configuracion_activa': configuracion_activa,
        'total_configuraciones': configuraciones.count()
    })

@login_required
@role_required('evaluador')
def crear_configuracion_email(request):
    """Vista para crear una nueva configuración de correo"""
    if request.method == 'POST':
        form = ConfiguracionEmailForm(request.POST)
        if form.is_valid():
            configuracion = form.save()
            messages.success(request, 'Configuración de correo creada exitosamente')
            return redirect('evaluacion:configuracion_email')
    else:
        form = ConfiguracionEmailForm()
    
    return render(request, 'evaluacion/configuracion_email_form.html', {
        'form': form,
        'titulo': 'Nueva Configuración de Email',
        'es_edicion': False
    })

@login_required
@role_required('evaluador')
def editar_configuracion_email(request, pk):
    """Vista para editar una configuración de correo existente"""
    configuracion = get_object_or_404(ConfiguracionEmail, pk=pk)
    
    if request.method == 'POST':
        form = ConfiguracionEmailForm(request.POST, instance=configuracion)
        if form.is_valid():
            form.save()
            messages.success(request, 'Configuración de correo actualizada exitosamente')
            return redirect('evaluacion:configuracion_email')
    else:
        form = ConfiguracionEmailForm(instance=configuracion)
    
    return render(request, 'evaluacion/configuracion_email_form.html', {
        'form': form,
        'configuracion': configuracion,
        'titulo': 'Editar Configuración de Email',
        'es_edicion': True
    })

@login_required
@role_required('evaluador')
def activar_configuracion_email(request, pk):
    """Vista AJAX para activar una configuración específica"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'})
    
    try:
        # Desactivar todas las configuraciones
        ConfiguracionEmail.objects.all().update(email_enabled=False)
        
        # Activar la seleccionada
        configuracion = get_object_or_404(ConfiguracionEmail, pk=pk)
        configuracion.email_enabled = True
        configuracion.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Configuración activada exitosamente',
            'configuracion_activa': {
                'id': configuracion.id,
                'email': configuracion.email_host_user,
                'tipo_proveedor': configuracion.get_tipo_proveedor_display()
            }
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
@role_required('evaluador')
def eliminar_configuracion_email(request, pk):
    """Vista para eliminar una configuración de correo"""
    configuracion = get_object_or_404(ConfiguracionEmail, pk=pk)
    
    if configuracion.email_enabled:
        messages.error(request, 'No se puede eliminar la configuración activa')
        return redirect('evaluacion:configuracion_email')
    
    if request.method == 'POST':
        configuracion.delete()
        messages.success(request, 'Configuración eliminada exitosamente')
        return redirect('evaluacion:configuracion_email')
    
    return render(request, 'evaluacion/confirmar_eliminacion_email.html', {
        'configuracion': configuracion
    })

@login_required
@role_required('evaluador')
def enviar_email_prueba(request):
    """Envía un email de prueba usando la configuración activa"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'})
    
    config = ConfiguracionEmail.get_configuracion_activa()
    if not config:
        return JsonResponse({'success': False, 'error': 'No hay una configuración de correo activa'})
    
    email_destino = request.POST.get('test_email')
    if not email_destino:
        return JsonResponse({'success': False, 'error': 'Debe proporcionar un email de destino'})
    
    try:
        success, message = config.enviar_email_prueba(email_destino)
        if success:
            return JsonResponse({
                'success': True,
                'message': f'Email de prueba enviado exitosamente a {email_destino}'
            })
        else:
            return JsonResponse({'success': False, 'error': message})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


# ============================================================================
# VISTAS DE DASHBOARD Y REPORTES - FASE 2.3
# ============================================================================

@login_required
@role_required('evaluador', 'supervisor', 'admin_tic', 'direccion')
def dashboard_rendimiento(request):
    """Dashboard con métricas y KPIs de rendimiento del sistema"""
    from accounts.models import User

    # Rango de fechas
    today = timezone.now().date()
    hace_30_dias = today - timedelta(days=30)
    primer_dia_mes = today.replace(day=1)

    # Estadísticas generales de solicitudes
    total_solicitudes = Solicitud.objects.count()
    solicitudes_mes = Solicitud.objects.filter(creada_el__date__gte=primer_dia_mes).count()
    solicitudes_30dias = Solicitud.objects.filter(creada_el__date__gte=hace_30_dias).count()

    pendientes = Solicitud.objects.filter(estado__in=['pendiente', 'en_revision']).count()
    aprobadas = Solicitud.objects.filter(estado='aprobada').count()
    rechazadas = Solicitud.objects.filter(estado='rechazada').count()

    # Estadísticas de evaluadores
    evaluadores = User.objects.filter(role='evaluador', activo=True)
    datos_evaluadores = []

    for evaluador in evaluadores:
        asignadas = Solicitud.objects.filter(evaluador_asignado=evaluador).count()
        evaluadas_mes = Solicitud.objects.filter(
            evaluador_asignado=evaluador,
            fecha_evaluacion__date__gte=primer_dia_mes,
            estado__in=['aprobada', 'rechazada']
        ).count()
        aprobadas_eval = Solicitud.objects.filter(
            evaluador_asignado=evaluador,
            estado='aprobada'
        ).count()

        datos_evaluadores.append({
            'nombre': evaluador.get_full_name() or evaluador.username,
            'asignadas': asignadas,
            'evaluadas_mes': evaluadas_mes,
            'aprobadas': aprobadas_eval,
            'porcentaje_aprobacion': round((aprobadas_eval / asignadas * 100), 1) if asignadas > 0 else 0
        })

    # Estadísticas de empresas
    total_empresas = Empresa.objects.count()
    empresas_activas = Empresa.objects.filter(activa=True).count()
    empresas_con_licencia = Empresa.objects.filter(numero_licencia__isnull=False).distinct().count()

    # Top 5 empresas con más solicitudes
    top_empresas = Empresa.objects.annotate(
        total_solicitudes=Count('solicitudes')
    ).order_by('-total_solicitudes')[:5]

    # Solicitudes por estado (para gráfico de pastel)
    solicitudes_por_estado = {
        'pendiente': Solicitud.objects.filter(estado='pendiente').count(),
        'en_revision': Solicitud.objects.filter(estado='en_revision').count(),
        'aprobada': Solicitud.objects.filter(estado='aprobada').count(),
        'rechazada': Solicitud.objects.filter(estado='rechazada').count(),
        'documentos_requeridos': Solicitud.objects.filter(estado='documentos_requeridos').count(),
    }

    # Solicitudes por mes (últimos 6 meses)
    solicitudes_por_mes = []
    for i in range(5, -1, -1):
        fecha = today - timedelta(days=30*i)
        primer_dia = fecha.replace(day=1)
        if i > 0:
            ultimo_dia = (fecha.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
        else:
            ultimo_dia = today

        count = Solicitud.objects.filter(
            creada_el__date__gte=primer_dia,
            creada_el__date__lte=ultimo_dia
        ).count()

        solicitudes_por_mes.append({
            'mes': primer_dia.strftime('%b %Y'),
            'count': count
        })

    # Calcular KPIs
    tasa_aprobacion = round((aprobadas / total_solicitudes * 100), 1) if total_solicitudes > 0 else 0
    tasa_rechazo = round((rechazadas / total_solicitudes * 100), 1) if total_solicitudes > 0 else 0

    # Tiempo promedio de evaluación (simulado - se puede calcular real)
    tiempo_promedio_evaluacion = 2.3  # días

    context = {
        'stats': {
            'total_solicitudes': total_solicitudes,
            'solicitudes_mes': solicitudes_mes,
            'solicitudes_30dias': solicitudes_30dias,
            'pendientes': pendientes,
            'aprobadas': aprobadas,
            'rechazadas': rechazadas,
            'tasa_aprobacion': tasa_aprobacion,
            'tasa_rechazo': tasa_rechazo,
            'tiempo_promedio': tiempo_promedio_evaluacion,
            'total_empresas': total_empresas,
            'empresas_activas': empresas_activas,
            'empresas_con_licencia': empresas_con_licencia,
        },
        'datos_evaluadores': datos_evaluadores,
        'top_empresas': top_empresas,
        'solicitudes_por_estado': solicitudes_por_estado,
        'solicitudes_por_mes': solicitudes_por_mes,
    }

    return render(request, 'evaluacion/dashboard/rendimiento.html', context)


@login_required
@role_required('evaluador', 'supervisor', 'admin_tic', 'direccion')
def distribucion_evaluadores(request):
    """Vista con gráfico de distribución de solicitudes por evaluador"""
    from accounts.models import User

    # Obtener todos los evaluadores activos
    evaluadores = User.objects.filter(role='evaluador', activo=True)

    datos_distribucion = []
    labels = []
    data_asignadas = []
    data_pendientes = []
    data_evaluadas = []

    for evaluador in evaluadores:
        nombre = evaluador.get_full_name() or evaluador.username
        asignadas = Solicitud.objects.filter(evaluador_asignado=evaluador).count()
        pendientes = Solicitud.objects.filter(
            evaluador_asignado=evaluador,
            estado__in=['pendiente', 'en_revision']
        ).count()
        evaluadas = Solicitud.objects.filter(
            evaluador_asignado=evaluador,
            estado__in=['aprobada', 'rechazada']
        ).count()

        labels.append(nombre)
        data_asignadas.append(asignadas)
        data_pendientes.append(pendientes)
        data_evaluadas.append(evaluadas)

        datos_distribucion.append({
            'evaluador': evaluador,
            'nombre': nombre,
            'asignadas': asignadas,
            'pendientes': pendientes,
            'evaluadas': evaluadas,
            'porcentaje_carga': round((asignadas / max(1, evaluadores.count())) * 100, 1)
        })

    # Calcular métricas de distribución
    total_asignadas = sum(data_asignadas)
    promedio_por_evaluador = round(total_asignadas / max(1, len(evaluadores)), 1)

    context = {
        'datos_distribucion': datos_distribucion,
        'labels': labels,
        'data_asignadas': data_asignadas,
        'data_pendientes': data_pendientes,
        'data_evaluadas': data_evaluadas,
        'total_asignadas': total_asignadas,
        'promedio_por_evaluador': promedio_por_evaluador,
        'total_evaluadores': evaluadores.count(),
    }

    return render(request, 'evaluacion/dashboard/distribucion_evaluadores.html', context)


@login_required
@role_required('evaluador', 'supervisor', 'admin_tic')
def historial_evaluaciones_empresa(request, empresa_id):
    """Vista de historial completo de evaluaciones de una empresa"""
    empresa = get_object_or_404(Empresa, id=empresa_id)

    # Obtener todas las solicitudes de la empresa
    solicitudes = Solicitud.objects.filter(empresa=empresa).select_related(
        'evaluador_asignado',
        'solicitante',
        'puerto_destino',
        'motivo_acceso'
    ).prefetch_related(
        'personal',
        'vehiculos'
    ).order_by('-creada_el')

    # Estadísticas de la empresa
    total_solicitudes = solicitudes.count()
    aprobadas = solicitudes.filter(estado='aprobada').count()
    rechazadas = solicitudes.filter(estado='rechazada').count()
    pendientes = solicitudes.filter(estado__in=['pendiente', 'en_revision']).count()

    tasa_aprobacion = round((aprobadas / total_solicitudes * 100), 1) if total_solicitudes > 0 else 0

    # Primera y última solicitud
    primera_solicitud = solicitudes.last()
    ultima_solicitud = solicitudes.first()

    # Paginación
    paginator = Paginator(solicitudes, 15)
    page = request.GET.get('page')
    solicitudes_page = paginator.get_page(page)

    context = {
        'empresa': empresa,
        'solicitudes': solicitudes_page,
        'stats': {
            'total': total_solicitudes,
            'aprobadas': aprobadas,
            'rechazadas': rechazadas,
            'pendientes': pendientes,
            'tasa_aprobacion': tasa_aprobacion,
        },
        'primera_solicitud': primera_solicitud,
        'ultima_solicitud': ultima_solicitud,
    }

    return render(request, 'evaluacion/dashboard/historial_empresa.html', context)


# ============================================================================
# EXPORTACIONES - FASE 2.3
# ============================================================================

@login_required
@role_required('evaluador', 'supervisor', 'admin_tic')
def exportar_empresas_csv(request):
    """Exporta el listado de empresas a formato CSV"""
    import csv
    from django.http import HttpResponse

    # Aplicar los mismos filtros que en gestionar_empresas
    busqueda = request.GET.get('q', '')
    estado_filtro = request.GET.get('estado', '')

    empresas = Empresa.objects.all().select_related('representante_legal')

    if busqueda:
        empresas = empresas.filter(
            Q(nombre__icontains=busqueda) |
            Q(rnc__icontains=busqueda) |
            Q(email__icontains=busqueda)
        )

    if estado_filtro:
        empresas = empresas.filter(estado=estado_filtro)

    # Crear respuesta HTTP con tipo CSV
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="empresas_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'

    # Agregar BOM para Excel
    response.write('\ufeff')

    writer = csv.writer(response)

    # Encabezados
    writer.writerow([
        'RNC',
        'Nombre',
        'Email',
        'Teléfono',
        'Dirección',
        'Estado',
        'Representante Legal',
        'Cédula Representante',
        'Fecha Registro',
        'Total Solicitudes',
        'Licencias Activas'
    ])

    # Datos
    for empresa in empresas:
        total_solicitudes = empresa.solicitudes.count()
        licencias_count = empresa.licencias_activas.count() if hasattr(empresa, 'licencias_activas') else 0

        writer.writerow([
            empresa.rnc,
            empresa.nombre,
            empresa.email or '',
            empresa.telefono or '',
            empresa.direccion or '',
            empresa.get_estado_display(),
            empresa.representante_legal.get_full_name() if empresa.representante_legal else '',
            empresa.representante_legal.cedula_rnc if empresa.representante_legal else '',
            empresa.created_at.strftime('%d/%m/%Y') if empresa.created_at else '',
            total_solicitudes,
            licencias_count
        ])

    return response


@login_required
@role_required('evaluador', 'supervisor', 'admin_tic')
def exportar_empresas_xlsx(request):
    """Exporta el listado de empresas a formato Excel (XLSX)"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from django.http import HttpResponse
    except ImportError:
        messages.error(request, 'La librería openpyxl no está instalada. Ejecuta: pip install openpyxl')
        return redirect('evaluacion:gestionar_empresas')

    # Aplicar filtros
    busqueda = request.GET.get('q', '')
    estado_filtro = request.GET.get('estado', '')

    empresas = Empresa.objects.all().select_related('representante_legal')

    if busqueda:
        empresas = empresas.filter(
            Q(nombre__icontains=busqueda) |
            Q(rnc__icontains=busqueda) |
            Q(email__icontains=busqueda)
        )

    if estado_filtro:
        empresas = empresas.filter(estado=estado_filtro)

    # Crear workbook y worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Empresas'

    # Estilos
    header_font = Font(bold=True, color='FFFFFF', size=12)
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Encabezados
    headers = [
        'RNC', 'Nombre', 'Email', 'Teléfono', 'Dirección', 'Estado',
        'Representante Legal', 'Cédula Representante', 'Fecha Registro',
        'Total Solicitudes', 'Licencias Activas'
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # Datos
    for row_num, empresa in enumerate(empresas, 2):
        total_solicitudes = empresa.solicitudes.count()
        licencias_count = empresa.licencias_activas.count() if hasattr(empresa, 'licencias_activas') else 0

        data = [
            empresa.rnc,
            empresa.nombre,
            empresa.email or '',
            empresa.telefono or '',
            empresa.direccion or '',
            empresa.get_estado_display(),
            empresa.representante_legal.get_full_name() if empresa.representante_legal else '',
            empresa.representante_legal.cedula_rnc if empresa.representante_legal else '',
            empresa.created_at.strftime('%d/%m/%Y') if empresa.created_at else '',
            total_solicitudes,
            licencias_count
        ]

        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            if col_num in [10, 11]:  # Columnas numéricas
                cell.alignment = Alignment(horizontal='center')

    # Ajustar ancho de columnas
    column_widths = [15, 35, 30, 15, 40, 15, 30, 18, 15, 18, 18]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # Crear respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="empresas_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'

    wb.save(response)
    return response


@login_required
@role_required('evaluador', 'supervisor', 'admin_tic')
def exportar_empresas_pdf(request):
    """Exporta el listado de empresas a formato PDF"""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from django.http import HttpResponse
        import io
    except ImportError:
        messages.error(request, 'La librería reportlab no está instalada. Ejecuta: pip install reportlab')
        return redirect('evaluacion:gestionar_empresas')

    # Aplicar filtros
    busqueda = request.GET.get('q', '')
    estado_filtro = request.GET.get('estado', '')

    empresas = Empresa.objects.all().select_related('representante_legal')

    if busqueda:
        empresas = empresas.filter(
            Q(nombre__icontains=busqueda) |
            Q(rnc__icontains=busqueda) |
            Q(email__icontains=busqueda)
        )

    if estado_filtro:
        empresas = empresas.filter(estado=estado_filtro)

    # Crear buffer
    buffer = io.BytesIO()

    # Crear PDF en orientación horizontal
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        rightMargin=30,
        leftMargin=30,
        topMargin=50,
        bottomMargin=30
    )

    # Container para elementos
    elements = []

    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#2c3e50'),
        spaceAfter=30,
        alignment=1  # Centrado
    )

    # Título
    title = Paragraph('Listado de Empresas', title_style)
    elements.append(title)

    # Fecha de generación
    fecha = Paragraph(
        f'Generado el: {timezone.now().strftime("%d/%m/%Y %H:%M")}',
        styles['Normal']
    )
    elements.append(fecha)
    elements.append(Spacer(1, 0.3*inch))

    # Datos para la tabla
    data = [['RNC', 'Nombre', 'Email', 'Estado', 'Representante', 'Solicitudes']]

    for empresa in empresas:
        total_solicitudes = empresa.solicitudes.count()
        data.append([
            empresa.rnc,
            empresa.nombre[:30] + '...' if len(empresa.nombre) > 30 else empresa.nombre,
            empresa.email[:25] + '...' if empresa.email and len(empresa.email) > 25 else (empresa.email or ''),
            empresa.get_estado_display(),
            empresa.representante_legal.get_full_name()[:25] if empresa.representante_legal else '',
            str(total_solicitudes)
        ])

    # Crear tabla
    table = Table(data, colWidths=[1.2*inch, 2.5*inch, 2*inch, 1*inch, 2*inch, 1*inch])

    # Estilo de la tabla
    table.setStyle(TableStyle([
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),

        # Datos
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),

        # Bordes
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
    ]))

    elements.append(table)

    # Construir PDF
    doc.build(elements)

    # Obtener valor del buffer
    pdf = buffer.getvalue()
    buffer.close()

    # Crear respuesta
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="empresas_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    response.write(pdf)

    return response
